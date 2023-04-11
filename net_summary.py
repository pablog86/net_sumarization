
import ipaddress

import tkinter as tk
from tkinter import filedialog
import os
from os import listdir
from os.path import isfile, join
import sys

import openpyxl

def sumarizar_subnets(subnets):
	ordered_subnets = sorted(subnets)
	print("Input:")
	print(ordered_subnets)
	
	flag = 1

	while flag == 1:
		summarized = set()
		flag = 0
		print("--------- is subnet lookup ---------")
		current_subnet = ordered_subnets[0]
		for i, subnet in enumerate(ordered_subnets[1:]):
			#print (current_subnet, subnet)
			if subnet.subnet_of(current_subnet) and subnet is not current_subnet:
				print(subnet, "is in", current_subnet)  
				ordered_subnets.remove(subnet) 		
				try:
					current_subnet = ordered_subnets[i+1]
				except IndexError as e:
					#pass
					print(e, i)
					print("items: ", ordered_subnets)
					print("lenght: ", len(ordered_subnets))
					#sys.exit()
			current_subnet = subnet
		print("--------- Contiguous lookup ---------")
		current_subnet = ordered_subnets[0]
		for subnet in ordered_subnets[1:]:
			
			if (subnet.network_address == current_subnet.broadcast_address + 1) and (subnet.prefixlen == current_subnet.prefixlen):
				#print (current_subnet, subnet)
				if subnet.supernet() == current_subnet.supernet():
					print("Contiguous networks: ", str(current_subnet), " <-> ", str(subnet))
					current_subnet = subnet.supernet()
					print("Supernet: ", str(current_subnet))
					summarized.add(current_subnet)
					flag = 1
				else:
					summarized.add(current_subnet)
					current_subnet = subnet
			else:
				summarized.add(current_subnet)
				current_subnet = subnet

		summarized.add(current_subnet)	#Last member
		ordered_subnets = summarized
		ordered_subnets = sorted(ordered_subnets)
		print("Qty of networks: ", len(ordered_subnets))
		print("Contiguous sumarization: ", ordered_subnets)
		print("")

	#summarized.add(current_subnet)	#Last member
	summarized = sorted(summarized)
	summarized = [str(address) for address in summarized]
	return list(summarized)



if __name__ == "__main__":

	root = tk.Tk()
	root.withdraw()
	path = filedialog.askopenfilename()
	wb = openpyxl.load_workbook(path)
	sheet = wb.active
	max_row = sheet.max_row

	subnets = []
	subnets = [c.value for c in sheet['A']]
	subnets_ip = [ipaddress.IPv4Network(subnet) for subnet in subnets]
	summary = sumarizar_subnets(subnets_ip)
	print("Total length of the input: ", len(subnets_ip))
	print("Summary length: ", len(summary))
	print("Output: ") 
	print(summary)
	print("List: ")

	for indice, net in enumerate(summary):
		print(net)
		celda = sheet.cell(row=indice+1, column=3)
		celda.value = net
	wb.save(path)
